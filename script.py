import requests
import datetime
import json
import openpyxl
#import smtplib
import os
from dotenv import load_dotenv 

load_dotenv() 
lookback_days = 1 # this is the number of days you would like to look back
SEEN_CVES_FILE = "seen_cves.json"
Vendor_list = []

def Get_vendors():
    # Load NVD data from a local .json file
    dataframe  = openpyxl.load_workbook('vendors_list.xlsx')
    
    # Define variable to read sheet
    dataframe1 = dataframe.active

    for row  in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            Vendor_list.append(col[row].value)
    return Vendor_list

def get_cvss_metrics(metrics):
    cvss_data = None
    vector_version = None

    if "cvssMetricV31" in metrics and metrics["cvssMetricV31"]:
        cvss_data = metrics["cvssMetricV31"][0]["cvssData"]

    elif "cvssMetricV40" in metrics and metrics["cvssMetricV40"]:
        cvss_data = metrics["cvssMetricV40"][0]["cvssData"]
    
    return cvss_data

def check_descriptions_language(description_data):
    description_info = None
    no_desc = "No description"
    
    for desc in description_data:
        
        if desc.get("lang") == "en":
            return desc.get("value", "")
    return no_desc


vendors_list = [v.strip() for v in Get_vendors() if v]

def cve_mentions_vendor(cve_entry):
    texts = []

    # Descriptions 
    for desc in cve_entry.get("descriptions", []):
        texts.append(desc.get("value", ""))

    # References (URLs or sources might contain vendor names)
    for ref in cve_entry.get("references", []):
        texts.append(ref.get("url", ""))
        texts.append(ref.get("source", ""))

    # (Optional) Metrics source
    metrics = cve_entry.get("metrics", {})
    for key in metrics:
        for metric_item in metrics[key]:
            texts.append(metric_item.get("source", ""))

    # Combine all text and search
    combined_text = " ".join(texts)
    combined_text_lower = combined_text.lower()
   
    mentioned_vendors = [vendor for vendor in vendors_list if vendor.lower() in combined_text_lower]
    
    return bool(mentioned_vendors)

def send_teams_alert(message_body, webhook_url):
    payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "summary": "CVE Alert",
        "themeColor": "0076D7",
        "title": " CVE Notification",
        "text": message_body.replace("\n", "<br>"),
        "sections": [
            {
                "text": "![funny gif](https://media3.giphy.com/media/v1.Y2lkPTc5MGI3NjExeHB4a3dldHdwbTJreXVwNDh2OWJiaW5mMmNya3J1cXo5dzl3YTBnYiZlcD12MV9pbnRlcm5hbF9naWZfYnlfaWQmY3Q9Zw/kFgzrTt798d2w/giphy.gif)"
            }
        ]
    }

    try:
        response = requests.post(
            webhook_url,
            data=json.dumps(payload),
            headers={"Content-Type": "application/json"}
        )
        if response.status_code == 200:
            print("Teams alert sent successfully.")
        else:
            print(f"Failed to send Teams alert. Status code: {response.status_code}")
    except Exception as e:
        print(f"Exception while sending Teams alert: {e}")


def fetch_recent_critical_cves_from_history():
   # Calculate UTC timestamps for now and 24 hours ago
    now = datetime.datetime.now(datetime.UTC)
    start_time = now - datetime.timedelta(days=lookback_days)

    # Format to required ISO 8601 format: YYYY-MM-DDTHH:MM:SSZ
    def format_iso8601_z(dt):
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

    start_str = format_iso8601_z(start_time)
    end_str = format_iso8601_z(now)

    # Fetch CVE history for that time range
    history_url = "https://services.nvd.nist.gov/rest/json/cves/2.0"
    params = {
        "pubStartDate": start_str,
        "pubEndDate": end_str
    }

    response = requests.get(history_url, params=params)
    
    if response.status_code != 200:
        print(f"Failed to retrive data {response.status_code}")
        exit()
        
    # Serializing json
    data = response.json()

    # The json object and writing to nvdcve-1.1-recent.json is for testing
    json_object = json.dumps(data, indent=4)
    
    # Writing to sample.json
    with open("nvdcve-1.1-recent.json", "w") as outfile:
       outfile.write(json_object)

    return data, now, start_str


def load_seen_cves():
    if os.path.exists(SEEN_CVES_FILE):
        with open(SEEN_CVES_FILE, "r") as file:
            return set(json.load(file))
    return set()

def save_seen_cves(cve_ids):
    with open(SEEN_CVES_FILE, "w") as file:
        json.dump(sorted(list(cve_ids)), file)

def sort_references(references_list):
    references_string = "References: "
    for ref in references_list:
        references_string = references_string + ref["url"] + "\n"

    return references_string

        
def main():
    text_body = []
    cve_count = 0

    # Fectching CVE data from NVD
    data, now, start_str = fetch_recent_critical_cves_from_history()

    totalResults = data["totalResults"]

    # Load seen CVEs ONCE
    seen_cves = load_seen_cves()
    new_cves_found = set()
    
    # Looping over the CVEs from NVD 
    for item in data.get("vulnerabilities", []):
        cve_data = item["cve"]
        cve_id = cve_data["id"]

        if cve_id in seen_cves:
            continue  # Skip cve already processed


        metrics = cve_data["metrics"]
        descriptions = cve_data["descriptions"]
        GetcvssMetric = get_cvss_metrics(metrics)
        
        
        if GetcvssMetric is None:
            continue
        

    # Getting critical vulnerabilities and then checking the vendors list
        baseSeverity = GetcvssMetric["baseSeverity"]
        if baseSeverity == "CRITICAL" :#and cve_mentions_vendor(cve_data):
   

            # Continue processing...
            new_cves_found.add(cve_id)
            # Update the seen CVE list
            
            NVD_url = "https://nvd.nist.gov/vuln/detail/"
            cve_id_str = str(cve_id)
            CVE_url = NVD_url+cve_id

            get_references = sort_references(cve_data["references"])
            
            
            cve_count += 1
            details = f"""CVE ID: {cve_id}
Published Date: {cve_data["published"]}
Modified Date: {cve_data["lastModified"]}
Severity: {baseSeverity}
Base Score: {GetcvssMetric["baseScore"]}
Vector: {GetcvssMetric["vectorString"]}
Description: {check_descriptions_language(descriptions)}
NVD Link: {CVE_url}
{get_references}

"""
            print(details)
            text_body.append(details)

    seen_cves.update(new_cves_found)
    save_seen_cves(seen_cves)


    # Add message to send based on relevant CVEs were found
    if cve_count == 0:
        message = f"Checking CVEs updated between {start_str} and {now}\n{totalResults} CVEs checked with no matches!!"
        print(message)
        text_body.append(message)
    if cve_count == 1:
        header = f"Checking CVEs updated between {start_str} and {now}\n{totalResults} CVEs checked you have {cve_count} CVE to look at!\n\n"
        print(header)
        text_body.insert(0, header)
    elif cve_count > 1:
        header = f"Checking CVEs updated between {start_str} and {now}\n{totalResults} CVEs checked you have {cve_count} CVEs to look at!\n\n"
        print(header)
        text_body.insert(0, header)

    # Send message alert
    send_teams_alert(
        message_body="".join(text_body),
        webhook_url=os.getenv("TEAMS_WEBHOOK")
    )


if __name__ == "__main__":
    main()


